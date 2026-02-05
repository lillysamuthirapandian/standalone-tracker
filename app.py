from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
from pathlib import Path
import fcntl
import time
from threading import Lock

app = Flask(__name__)
CORS(app, resources={
    r"/*": {
        "origins": [
            "https://teams.microsoft.com",
            "https://*.teams.microsoft.com",
            "https://*.office.com",
            "*"  # Allow all for development (remove in production)
        ]
    }
})

# Thread lock for file operations
file_lock = Lock()

# File paths
MASTER_FILE = 'tracker_master_data.xlsx'
GENERATED_FILE = 'tracker_generated_report.xlsx'
USERS_FILE = 'users_master_data.xlsx'

# Field mapping for better readability
FIELD_LABELS = {
    'id': 'ID',
    'originalId': 'Original ID',
    'lastModified': 'Last Modified',
    'customer': 'Customer',
    'servicePackVersion': 'Service Pack Version',
    'dateOfShipment': 'Date of Shipment',
    'saveFileLibrary': 'Save File Library',
    'saveFileName': 'Save File Name',
    'salesforceId': 'Salesforce ID',
    'jiraId': 'Jira ID',
    'issueDescription': 'Issue Description',
    'objectName': 'Object Name',
    'objectType': 'Object Type',
    'objectDescriptionWithVersion': 'Object Description with Version',
    'actionType': 'Action Type',
    'downtimeRequired': 'Downtime Required',
    'specialInstructions': 'Special Instructions',
    'shippedBy': 'Shipped By',
    'vendorName': 'Vendor Name',
    'destinationObjectLibraryType': 'Destination Object Library Type',
    'fileTransferLink': 'File Transfer Link'
}

ALL_FIELDS = list(FIELD_LABELS.keys())


def acquire_file_lock(file_handle):
    """Acquire an exclusive lock on a file"""
    max_retries = 10
    retry_count = 0
    while retry_count < max_retries:
        try:
            fcntl.flock(file_handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            return True
        except IOError:
            retry_count += 1
            time.sleep(0.1)  # Wait 100ms before retry
    return False


def release_file_lock(file_handle):
    """Release the lock on a file"""
    try:
        fcntl.flock(file_handle.fileno(), fcntl.LOCK_UN)
    except:
        pass


def initialize_users_file():
    """Initialize the users Excel file if it doesn't exist"""
    if not os.path.exists(USERS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Users'
        
        # Add headers with styling
        headers = ['ID', 'Name', 'Short Name', 'Employee ID']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
        wb.save(USERS_FILE)
        wb.close()


def initialize_master_file():
    """Initialize the master Excel file if it doesn't exist"""
    if not os.path.exists(MASTER_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Tracker Data'
        
        # Add headers with styling
        headers = ['ID'] + [FIELD_LABELS[field] for field in ALL_FIELDS[1:]]
        ws.append(headers)
        
        # Style the header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        column_widths = {
            'A': 8, 'B': 20, 'C': 15, 'D': 20, 'E': 20, 'F': 15, 'G': 15,
            'H': 30, 'I': 20, 'J': 15, 'K': 30, 'L': 15, 'M': 15, 'N': 30,
            'O': 15, 'P': 15, 'Q': 25, 'R': 30
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        wb.save(MASTER_FILE)


def get_next_id():
    """Get the next available ID"""
    wb = openpyxl.load_workbook(MASTER_FILE)
    ws = wb.active
    
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            try:
                current_id = int(row[0])
                max_id = max(max_id, current_id)
            except (ValueError, TypeError):
                pass
    
    wb.close()
    return max_id + 1


def read_all_data():
    """Read all data from master file and renumber IDs sequentially"""
    initialize_master_file()
    
    # Use thread lock for read
    with file_lock:
        lock_file = open(MASTER_FILE, 'r')
        if not acquire_file_lock(lock_file):
            lock_file.close()
            raise Exception("Could not acquire file lock. Please try again.")
        
        try:
            wb = openpyxl.load_workbook(MASTER_FILE)
            ws = wb.active
            
            data = []
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:  # Check if ID exists
                    record = {}
                    for i, header in enumerate(headers):
                        # Convert header back to field key
                        field_key = None
                        for key, label in FIELD_LABELS.items():
                            if label == header:
                                field_key = key
                                break
                        
                        if field_key:
                            value = row[i] if i < len(row) else None
                            # Convert date objects to string
                            if isinstance(value, datetime):
                                value = value.strftime('%Y-%m-%d')
                            record[field_key] = value
                    
                    data.append(record)
            
            wb.close()
        finally:
            release_file_lock(lock_file)
            lock_file.close()
    
    # Renumber IDs sequentially while preserving original IDs
    for idx, record in enumerate(data, start=1):
        # Store original ID if not already stored
        if not record.get('originalId'):
            record['originalId'] = record['id']
        # Renumber to sequential ID
        record['id'] = idx
    
    return data


def write_data_to_master(data):
    """Write data to master file with file locking"""
    initialize_master_file()
    
    # Use thread lock and file lock
    with file_lock:
        # Open file for locking
        lock_file = open(MASTER_FILE, 'r+')
        if not acquire_file_lock(lock_file):
            lock_file.close()
            raise Exception("Could not acquire file lock. Please try again.")
        
        try:
            wb = openpyxl.load_workbook(MASTER_FILE)
            ws = wb.active
            
            # Clear existing data (keep headers)
            ws.delete_rows(2, ws.max_row)
            
            # Write data
            for record in data:
                row_data = []
                for field in ALL_FIELDS:
                    value = record.get(field, '')
                    row_data.append(value)
                ws.append(row_data)
            
            wb.save(MASTER_FILE)
            wb.close()
        finally:
            release_file_lock(lock_file)
            lock_file.close()


@app.route('/')
def index():
    """Serve the index.html file"""
    return send_file('index.html')


@app.route('/config.html')
def config():
    """Serve the Teams configuration page"""
    return send_file('teams/config.html')


# ============ USER MANAGEMENT ROUTES ============

@app.route('/users', methods=['GET'])
def get_users():
    """Get all users"""
    try:
        initialize_users_file()
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active
        
        users = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If ID exists
                user = {
                    'id': row[0],
                    'name': row[1] or '',
                    'shortName': row[2] or '',
                    'employeeId': row[3] or ''
                }
                users.append(user)
        
        wb.close()
        return jsonify(users)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/users/add', methods=['POST'])
def add_user():
    """Add a new user"""
    try:
        data = request.json
        initialize_users_file()
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active
        
        # Generate new ID
        max_id = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                max_id = max(max_id, int(row[0]))
        new_id = max_id + 1
        
        # Add new user
        ws.append([
            new_id,
            data.get('name', ''),
            data.get('shortName', ''),
            data.get('employeeId', '')
        ])
        
        wb.save(USERS_FILE)
        wb.close()
        
        return jsonify({'success': True, 'id': new_id, 'message': 'User added successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/users/update/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    """Update an existing user"""
    try:
        data = request.json
        initialize_users_file()
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active
        
        # Find and update the user
        for row in ws.iter_rows(min_row=2):
            if row[0].value == user_id:
                row[1].value = data.get('name', '')
                row[2].value = data.get('shortName', '')
                row[3].value = data.get('employeeId', '')
                break
        
        wb.save(USERS_FILE)
        wb.close()
        
        return jsonify({'success': True, 'message': 'User updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/users/delete/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    """Delete a user"""
    try:
        initialize_users_file()
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active
        
        # Find and delete the user
        row_to_delete = None
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == user_id:
                row_to_delete = idx
                break
        
        if row_to_delete:
            ws.delete_rows(row_to_delete, 1)
            wb.save(USERS_FILE)
        
        wb.close()
        return jsonify({'success': True, 'message': 'User deleted successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============ TRACKER ENTRY ROUTES ============

@app.route('/add', methods=['POST'])
def add_entry():
    """Add a new entry"""
    try:
        data = request.json
        data['id'] = get_next_id()
        data['lastModified'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        all_data = read_all_data()
        all_data.append(data)
        write_data_to_master(all_data)
        
        return jsonify({'message': 'Entry added successfully', 'id': data['id']}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/update/<int:record_id>', methods=['PUT'])
def update_entry(record_id):
    """Update an existing entry with concurrency check"""
    try:
        updated_data = request.json
        client_timestamp = updated_data.get('lastModified', '')
        all_data = read_all_data()
        
        found = False
        for i, record in enumerate(all_data):
            if int(record.get('id', 0)) == record_id:
                # Check if record was modified by another user
                current_timestamp = record.get('lastModified', '')
                if client_timestamp and current_timestamp and client_timestamp != current_timestamp:
                    return jsonify({
                        'error': 'CONFLICT',
                        'message': 'This record was modified by another user. Please refresh and try again.',
                        'current_data': record
                    }), 409
                
                # Update with new timestamp
                updated_data['id'] = record_id
                updated_data['lastModified'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                all_data[i] = updated_data
                found = True
                break
        
        if not found:
            return jsonify({'error': 'Record not found'}), 404
        
        write_data_to_master(all_data)
        return jsonify({'message': 'Entry updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/delete/<int:record_id>', methods=['DELETE'])
def delete_entry(record_id):
    """Delete an entry"""
    try:
        all_data = read_all_data()
        original_length = len(all_data)
        
        all_data = [record for record in all_data if int(record.get('id', 0)) != record_id]
        
        if len(all_data) == original_length:
            return jsonify({'error': 'Record not found'}), 404
        
        write_data_to_master(all_data)
        return jsonify({'message': 'Entry deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/data', methods=['GET'])
def get_data():
    """Get all data"""
    try:
        data = read_all_data()
        return jsonify(data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/generate', methods=['POST'])
def generate_excel():
    """Generate filtered Excel report with selected fields"""
    try:
        selected_fields = request.json.get('fields', ALL_FIELDS)
        
        # Always include ID as the first field
        if 'id' not in selected_fields:
            selected_fields.insert(0, 'id')
        
        all_data = read_all_data()
        
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Standalone Tracker'
        
        # Add headers
        headers = [FIELD_LABELS[field] for field in selected_fields]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        header_font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add data rows
        data_font = Font(name='Calibri', size=11)
        for record in all_data:
            row_data = []
            for field in selected_fields:
                value = record.get(field, '')
                row_data.append(value)
            ws.append(row_data)
        
        # Style data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = data_font
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Save the file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Standalone_Tracker_{timestamp}.xlsx'
        wb.save(filename)
        wb.close()
        
        return send_file(filename, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/generate-selected', methods=['POST'])
def generate_selected_excel():
    """Generate Excel for selected entries with header and detail sections"""
    try:
        selected_ids = request.json.get('ids', [])
        
        if not selected_ids:
            return jsonify({'error': 'No IDs provided'}), 400
        
        all_data = read_all_data()
        selected_data = [record for record in all_data if int(record.get('id', 0)) in selected_ids]
        
        if not selected_data:
            return jsonify({'error': 'No records found'}), 404
        
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Code Delivery Sheet'
        
        # Styling - Blue theme
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        section_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        section_font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
        data_font = Font(name='Calibri', size=11)
        
        current_row = 1
        
        for record in selected_data:
            # Header fields
            header_fields = [
                ('vendorName', 'Vendor Name'),
                ('customer', 'Customer'),
                ('servicePackVersion', 'Service Pack Version'),
                ('dateOfShipment', 'Date of Shipment'),
                ('saveFileLibrary', 'Save File Library'),
                ('saveFileName', 'Save File Name'),
                ('shippedBy', 'Shipped By'),
                ('salesforceId', 'Salesforce ID'),
                ('jiraId', 'Jira ID'),
                ('fileTransferLink', 'File Transfer Link'),
                ('issueDescription', 'Issue Description')
            ]
            
            for field_key, field_label in header_fields:
                ws.cell(row=current_row, column=1, value=field_label).font = Font(name='Calibri', size=11, bold=True)
                ws.cell(row=current_row, column=2, value=record.get(field_key, '')).font = data_font
                current_row += 1
            
            current_row += 1
            
            # Detail headers
            detail_headers = ['Object Name', 'Object Type', 'Object Description', 'Action Type', 'Destination Type', 'Downtime', 'Special Instructions']
            for col, header in enumerate(detail_headers, start=1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            current_row += 1
            
            # Parse detail lines (split by pipe)
            object_names = (record.get('objectName', '') or '').split(' | ')
            object_types = (record.get('objectType', '') or '').split(' | ')
            object_descs = (record.get('objectDescriptionWithVersion', '') or '').split(' | ')
            action_types = (record.get('actionType', '') or '').split(' | ')
            dest_types = (record.get('destinationObjectLibraryType', '') or '').split(' | ')
            downtimes = (record.get('downtimeRequired', '') or '').split(' | ')
            instructions = (record.get('specialInstructions', '') or '').split(' | ')
            
            max_details = max(len(object_names), len(object_types), len(object_descs), 
                            len(action_types), len(dest_types), len(downtimes), len(instructions))
            
            for i in range(max_details):
                ws.cell(row=current_row, column=1, value=object_names[i] if i < len(object_names) else '').font = data_font
                ws.cell(row=current_row, column=2, value=object_types[i] if i < len(object_types) else '').font = data_font
                ws.cell(row=current_row, column=3, value=object_descs[i] if i < len(object_descs) else '').font = data_font
                ws.cell(row=current_row, column=4, value=action_types[i] if i < len(action_types) else '').font = data_font
                ws.cell(row=current_row, column=5, value=dest_types[i] if i < len(dest_types) else '').font = data_font
                ws.cell(row=current_row, column=6, value=downtimes[i] if i < len(downtimes) else '').font = data_font
                ws.cell(row=current_row, column=7, value=instructions[i] if i < len(instructions) else '').font = data_font
                
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).alignment = Alignment(vertical='center', wrap_text=True)
                
                current_row += 1
            
            current_row += 2  # Space between entries
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 40
        
        # Save the file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'tracker_entries_{timestamp}.xlsx'
        wb.save(filename)
        wb.close()
        
        return send_file(filename, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    initialize_master_file()
    initialize_users_file()
    app.run(debug=True, host='0.0.0.0', port=5000)
