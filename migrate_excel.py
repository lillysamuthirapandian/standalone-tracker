import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Load existing file
wb = openpyxl.load_workbook('tracker_master_data.xlsx')
ws = wb.active

# Read all existing data (skip header)
existing_data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:  # If ID exists
        existing_data.append(list(row))

print(f"Found {len(existing_data)} existing entries")

# Clear worksheet
ws.delete_rows(1, ws.max_row)

# Add new headers with all fields
new_headers = ['ID', 'Original ID', 'Last Modified', 'Customer', 'Service Pack Version', 
               'Date of Shipment', 'Save File Library', 'Save File Name', 'Salesforce ID', 
               'Jira ID', 'Issue Description', 'Object Name', 'Object Type', 
               'Object Description with Version', 'Action Type', 'Downtime Required', 
               'Special Instructions', 'Shipped By', 'Vendor Name', 
               'Destination Object Library Type', 'File Transfer Link']

ws.append(new_headers)

# Style headers
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Migrate existing data to new structure
for old_row in existing_data:
    new_row = [
        old_row[0] if len(old_row) > 0 else '',  # ID
        old_row[0] if len(old_row) > 0 else '',  # Original ID (same as ID initially)
        '',  # Last Modified (empty for existing data)
        old_row[1] if len(old_row) > 1 else '',  # Customer
        '',  # Service Pack Version (new field, empty)
        old_row[2] if len(old_row) > 2 else '',  # Date of Shipment
        old_row[3] if len(old_row) > 3 else '',  # Save File Library
        old_row[4] if len(old_row) > 4 else '',  # Save File Name
        old_row[5] if len(old_row) > 5 else '',  # Salesforce ID
        old_row[6] if len(old_row) > 6 else '',  # Jira ID
        old_row[7] if len(old_row) > 7 else '',  # Issue Description
        old_row[8] if len(old_row) > 8 else '',  # Object Name
        old_row[9] if len(old_row) > 9 else '',  # Object Type
        old_row[10] if len(old_row) > 10 else '',  # Object Description with Version
        old_row[11] if len(old_row) > 11 else '',  # Action Type
        old_row[12] if len(old_row) > 12 else '',  # Downtime Required
        old_row[13] if len(old_row) > 13 else '',  # Special Instructions
        old_row[14] if len(old_row) > 14 else '',  # Shipped By
        old_row[15] if len(old_row) > 15 else '',  # Vendor Name
        old_row[16] if len(old_row) > 16 else '',  # Destination Object Library Type
        old_row[17] if len(old_row) > 17 else '',  # File Transfer Link
    ]
    ws.append(new_row)

wb.save('tracker_master_data.xlsx')
wb.close()
print(f"Migration complete! {len(existing_data)} entries preserved with new structure.")
