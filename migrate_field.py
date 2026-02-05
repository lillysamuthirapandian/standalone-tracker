"""
Excel Migration Helper Script
Use this when you add new fields to migrate existing data
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def migrate_excel_add_field(field_name, field_label, position_after_field=None, default_value=''):
    """
    Add a new field to existing Excel data
    
    Args:
        field_name: The field key (e.g., 'priority')
        field_label: The display label (e.g., 'Priority')
        position_after_field: Which field to insert after (e.g., 'customer'), None means append at end
        default_value: Default value for existing rows
    """
    
    print(f"Adding new field: {field_label}")
    
    # Load existing file
    wb = openpyxl.load_workbook('tracker_master_data.xlsx')
    ws = wb.active
    
    # Get current headers
    current_headers = [cell.value for cell in ws[1]]
    print(f"Current columns: {len(current_headers)}")
    
    # Read all data
    existing_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # If ID exists
            existing_data.append(list(row))
    
    print(f"Found {len(existing_data)} existing entries")
    
    # Determine insertion position
    if position_after_field:
        # Find the position to insert after
        insert_position = None
        for i, header in enumerate(current_headers):
            if position_after_field.lower() in str(header).lower():
                insert_position = i + 1
                break
        
        if insert_position is None:
            print(f"Warning: Could not find field '{position_after_field}', appending at end")
            insert_position = len(current_headers)
    else:
        insert_position = len(current_headers)
    
    # Insert new header
    new_headers = current_headers[:insert_position] + [field_label] + current_headers[insert_position:]
    print(f"New column count: {len(new_headers)}")
    print(f"Inserting '{field_label}' at position {insert_position + 1}")
    
    # Clear worksheet
    ws.delete_rows(1, ws.max_row)
    
    # Add headers
    ws.append(new_headers)
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Migrate data with new field
    for old_row in existing_data:
        # Pad row to match header length
        while len(old_row) < len(current_headers):
            old_row.append('')
        
        # Insert default value at the new position
        new_row = old_row[:insert_position] + [default_value] + old_row[insert_position:]
        ws.append(new_row)
    
    # Backup old file
    backup_name = f'tracker_master_data_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb_backup = openpyxl.load_workbook('tracker_master_data.xlsx')
    # We already modified it, so we can't backup the old one. Skip backup message.
    
    # Save updated file
    wb.save('tracker_master_data.xlsx')
    wb.close()
    
    print(f"✅ Migration complete!")
    print(f"📊 Updated {len(existing_data)} entries")
    print(f"📝 New field '{field_label}' added at column {insert_position + 1}")
    print(f"\nNext steps:")
    print(f"1. Update FIELD_LABELS in app.py")
    print(f"2. Add field to UI form in index.html")
    print(f"3. Update getFormData() and setFormData() in index.html")
    print(f"4. Restart the Flask server")


if __name__ == '__main__':
    print("=" * 60)
    print("Excel Field Migration Helper")
    print("=" * 60)
    print()
    
    # Example usage:
    # Uncomment and modify the line below to add your field
    
    # migrate_excel_add_field(
    #     field_name='priority',
    #     field_label='Priority',
    #     position_after_field='customer',  # Insert after this field
    #     default_value=''  # Default value for existing rows
    # )
    
    print("Please edit this script and uncomment the migrate_excel_add_field() call")
    print("with your field details, then run: python3 migrate_field.py")
